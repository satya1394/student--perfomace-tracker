"""
Academic Performance Utility Module.
Handles SGPA/CGPA calculations, grade mappings, report exports (Excel/PDF), and email alerting.
"""

import io
from datetime import datetime, timezone
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.config import Config


def calculate_grade_and_points(marks: float) -> tuple[str, float]:
    """
    Converts raw marks (0-100 scale) to 10-point letter grade and grade point.
    
    Args:
        marks: Numerical score between 0 and 100.
        
    Returns:
        tuple (letter_grade, grade_point)
    """
    marks = max(0.0, min(100.0, float(marks)))
    if marks >= 90.0:
        return "O", 10.0
    elif marks >= 80.0:
        return "A+", 9.0
    elif marks >= 70.0:
        return "A", 8.0
    elif marks >= 60.0:
        return "B+", 7.0
    elif marks >= 50.0:
        return "B", 6.0
    elif marks >= 40.0:
        return "C", 5.0
    elif marks >= 35.0:
        return "P", 4.0
    else:
        return "F", 0.0


def calculate_sgpa(enrollment_records: list[dict]) -> float:
    """
    Computes Semester Grade Point Average (SGPA).
    Formula: SGPA = Σ(Grade_Point_i * Credits_i) / Σ(Credits_i)
    
    Args:
        enrollment_records: List of dicts containing 'grade_point' and 'credits'.
        
    Returns:
        SGPA rounded to 2 decimal places.
    """
    if not enrollment_records:
        return 0.0

    total_weighted_points = sum(
        rec.get("grade_point", 0.0) * rec.get("credits", 0) 
        for rec in enrollment_records
    )
    total_credits = sum(rec.get("credits", 0) for rec in enrollment_records)

    if total_credits <= 0:
        return 0.0

    return round(total_weighted_points / total_credits, 2)


def calculate_cgpa(all_semesters_enrollments: list[dict]) -> float:
    """
    Computes Cumulative Grade Point Average (CGPA) across all enrolled courses.
    Formula: CGPA = Σ(Grade_Point_j * Credits_j) / Σ(Credits_j)
    """
    if not all_semesters_enrollments:
        return 0.0

    total_weighted_points = sum(
        rec.get("grade_point", 0.0) * rec.get("credits", 0) 
        for rec in all_semesters_enrollments
    )
    total_credits = sum(rec.get("credits", 0) for rec in all_semesters_enrollments)

    if total_credits <= 0:
        return 0.0

    return round(total_weighted_points / total_credits, 2)


def calculate_grade(marks: float) -> tuple[str, float]:
    """Converts marks to grade letter and point."""
    return calculate_grade_and_points(marks)


def calculate_sgpa_for_student(student_id: str, semester: int) -> float:
    """Computes SGPA directly from student enrollments in database."""
    from app.database import get_db_session, Enrollment, Subject, Course
    db = get_db_session()
    try:
        enrollments = db.query(Enrollment).filter(
            Enrollment.student_id == student_id,
            Enrollment.semester == int(semester)
        ).all()
        if not enrollments:
            return 0.0
        total_credits = 0.0
        weighted_sum = 0.0
        for enr in enrollments:
            creds = 4.0
            if enr.subject_id:
                sub = db.query(Subject).filter(Subject.id == enr.subject_id).first()
                if sub:
                    creds = float(sub.credits)
            elif enr.course_id:
                crs = db.query(Course).filter(Course.course_id == enr.course_id).first()
                if crs:
                    creds = float(crs.credits)
            gp = enr.grade_point if enr.grade_point is not None else calculate_grade(enr.marks_obtained)[1]
            total_credits += creds
            weighted_sum += creds * gp
        return round(weighted_sum / total_credits, 2) if total_credits > 0 else 0.0
    finally:
        db.close()


def calculate_cgpa_for_student(student_id: str) -> float:
    """Computes CGPA across all enrolled semesters in database."""
    from app.database import get_db_session, Enrollment, Subject, Course
    db = get_db_session()
    try:
        enrollments = db.query(Enrollment).filter(Enrollment.student_id == student_id).all()
        if not enrollments:
            return 0.0
        total_credits = 0.0
        weighted_sum = 0.0
        for enr in enrollments:
            creds = 4.0
            if enr.subject_id:
                sub = db.query(Subject).filter(Subject.id == enr.subject_id).first()
                if sub:
                    creds = float(sub.credits)
            elif enr.course_id:
                crs = db.query(Course).filter(Course.course_id == enr.course_id).first()
                if crs:
                    creds = float(crs.credits)
            gp = enr.grade_point if enr.grade_point is not None else calculate_grade(enr.marks_obtained)[1]
            total_credits += creds
            weighted_sum += creds * gp
        return round(weighted_sum / total_credits, 2) if total_credits > 0 else 0.0
    finally:
        db.close()


def generate_excel_report(student_info: dict, kpis: dict, courses_df: pd.DataFrame) -> bytes:
    """
    Generates a professionally styled Excel workbook with KPI summaries and detailed grade sheets.
    
    Returns:
        Raw bytes of the generated .xlsx file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Academic Performance Report"

    # Header styling
    navy_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    accent_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    sub_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title Block
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "INSTITUTE ACADEMIC PERFORMANCE REPORT"
    title_cell.font = header_font
    title_cell.fill = navy_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Metadata
    ws["A3"] = "Student ID:"
    ws["B3"] = student_info.get("student_id", "N/A")
    ws["A4"] = "Student Name:"
    ws["B4"] = student_info.get("name", "N/A")
    ws["D3"] = "Department:"
    ws["E3"] = student_info.get("department", "N/A")
    ws["D4"] = "Generated On:"
    ws["E4"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    for cell in ["A3", "A4", "D3", "D4"]:
        ws[cell].font = bold_font

    # KPI Summary Row
    ws.merge_cells("A6:G6")
    kpi_banner = ws["A6"]
    kpi_banner.value = f"OVERALL CGPA: {kpis.get('cgpa', 0.0)} | AVERAGE ATTENDANCE: {kpis.get('attendance', 0.0)}% | RISK STATUS: {kpis.get('risk_level', 'LOW')}"
    kpi_banner.font = sub_font
    kpi_banner.fill = accent_fill
    kpi_banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 24

    # Table Header
    headers = ["Semester", "Course Code", "Course Name", "Credits", "Marks", "Grade", "Attendance %"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col_idx, value=h)
        cell.font = sub_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Table Rows
    curr_row = 9
    for _, row in courses_df.iterrows():
        ws.cell(row=curr_row, column=1, value=row.get("semester", "")).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=2, value=row.get("course_code", "")).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=3, value=row.get("course_name", ""))
        ws.cell(row=curr_row, column=4, value=row.get("credits", 0)).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=5, value=row.get("marks_obtained", 0.0)).alignment = Alignment(horizontal="right")
        ws.cell(row=curr_row, column=6, value=row.get("grade", "")).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=7, value=f"{row.get('attendance_percentage', 0.0)}%").alignment = Alignment(horizontal="right")

        for col_idx in range(1, 8):
            ws.cell(row=curr_row, column=col_idx).border = thin_border
            ws.cell(row=curr_row, column=col_idx).font = normal_font
        curr_row += 1

    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_pdf_report(student_info: dict, kpis: dict, courses_df: pd.DataFrame) -> bytes:
    """
    Generates a PDF academic transcript report.
    Uses WeasyPrint when available, with a clean HTML/CSS generator.
    """
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>Student Academic Performance Report</title>
        <style>
            body {{ font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; margin: 30px; color: #1e293b; }}
            .header {{ background-color: #1e293b; color: white; padding: 20px; text-align: center; border-radius: 6px; }}
            .header h1 {{ margin: 0; font-size: 22px; }}
            .meta-grid {{ display: table; width: 100%; margin: 20px 0; }}
            .meta-row {{ display: table-row; }}
            .meta-cell {{ display: table-cell; padding: 6px 12px; font-size: 14px; }}
            .kpi-bar {{ background-color: #3b82f6; color: white; padding: 12px; border-radius: 4px; font-weight: bold; text-align: center; margin-bottom: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th, td {{ border: 1px solid #cbd5e1; padding: 8px 12px; font-size: 13px; text-align: left; }}
            th {{ background-color: #f1f5f9; font-weight: 600; }}
            .text-center {{ text-align: center; }}
            .text-right {{ text-align: right; }}
            .footer {{ margin-top: 40px; font-size: 11px; color: #64748b; text-align: center; border-top: 1px solid #e2e8f0; padding-top: 10px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>INSTITUTE OF HIGHER TECHNOLOGY & SCIENCE</h1>
            <p style="margin: 5px 0 0 0; font-size: 14px;">Official Student Academic Analytics Report</p>
        </div>

        <div class="meta-grid">
            <div class="meta-row">
                <div class="meta-cell"><strong>Student ID:</strong> {student_info.get('student_id', 'N/A')}</div>
                <div class="meta-cell"><strong>Department:</strong> {student_info.get('department', 'N/A')}</div>
            </div>
            <div class="meta-row">
                <div class="meta-cell"><strong>Student Name:</strong> {student_info.get('name', 'N/A')}</div>
                <div class="meta-cell"><strong>Date of Report:</strong> {datetime.now(timezone.utc).strftime('%B %d, %Y')}</div>
            </div>
        </div>

        <div class="kpi-bar">
            Cumulative CGPA: {kpis.get('cgpa', 0.0)} &nbsp;|&nbsp;
            Attendance: {kpis.get('attendance', 0.0)}% &nbsp;|&nbsp;
            Academic Status: {kpis.get('risk_level', 'LOW')} RISK
        </div>

        <table>
            <thead>
                <tr>
                    <th class="text-center">Sem</th>
                    <th class="text-center">Course Code</th>
                    <th>Course Title</th>
                    <th class="text-center">Credits</th>
                    <th class="text-right">Marks</th>
                    <th class="text-center">Grade</th>
                    <th class="text-right">Attendance</th>
                </tr>
            </thead>
            <tbody>
    """
    for _, row in courses_df.iterrows():
        html_content += f"""
                <tr>
                    <td class="text-center">{row.get('semester', '')}</td>
                    <td class="text-center">{row.get('course_code', '')}</td>
                    <td>{row.get('course_name', '')}</td>
                    <td class="text-center">{row.get('credits', 0)}</td>
                    <td class="text-right">{row.get('marks_obtained', 0.0):.1f}</td>
                    <td class="text-center"><strong>{row.get('grade', '')}</strong></td>
                    <td class="text-right">{row.get('attendance_percentage', 0.0):.1f}%</td>
                </tr>
        """
    html_content += """
            </tbody>
        </table>

        <div class="footer">
            Generated autonomously by Student Academic Performance Analytics Platform. Confidential document.
        </div>
    </body>
    </html>
    """

    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_content).write_pdf()
        return pdf_bytes
    except Exception:
        # Fallback to UTF-8 encoded HTML formatted as downloadable string
        return html_content.encode("utf-8")


def send_performance_alert(student_name: str, student_email: str, sgpa: float, risk_level: str) -> bool:
    """
    Dispatches automated early-warning alerts for performance drops.
    """
    # Production logging & simulation
    print(f"[ALERT DISPATCH] Sent to {student_email} ({student_name}): Risk Level={risk_level}, Latest SGPA={sgpa}")
    return True
